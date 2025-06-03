import streamlit as st
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
import os
import unicodedata

DATA_FILE = "waste_log.xlsx"
PRODUCTS_FILE = "products.xlsx"

# ──────────────────────────
# 商品リスト読み込み（単価列は無視）
# ──────────────────────────
@st.cache_data
def load_products():
    try:
        return pd.read_excel(PRODUCTS_FILE, engine="openpyxl")[["商品名"]]
    except FileNotFoundError:
        st.error(f"商品データ {PRODUCTS_FILE} が見つかりません")
        return pd.DataFrame(columns=["商品名"])

products_df = load_products()

# ──────────────────────────
# ユーティリティ
# ──────────────────────────
def normalize_text(text: str) -> str:
    text = unicodedata.normalize("NFKC", text).lower()
    text = text.translate(str.maketrans(
        "アイウエオカキクケコサシスセソタチツテトナニヌネノ"
        "ハヒフヘホマミムメモヤユヨラリルレロワヲン",
        "あいうえおかきくけこさしすせそたちつてとなにぬねの"
        "はひふへほまみむめもやゆよらりるれろわをん"
    ))
    return text

def set_column_width(sheet, num_days: int, width: int = 12) -> None:
    for i in range(num_days + 2):                  # 商品名 + 日数 + 合計列
        sheet.column_dimensions[get_column_letter(i + 1)].width = width

# ──────────────────────────
# Streamlit UI
# ──────────────────────────
st.title("🐟 fishパンロス")

# 1. 商品名フィルタ
input_text = st.text_input("商品の頭文字を入力してください").strip()

if input_text:
    key = normalize_text(input_text)
    cand_df = products_df[products_df["商品名"].apply(
        lambda x: key in normalize_text(str(x))
    )]
    if cand_df.empty:
        st.warning("該当する商品が見つかりません")
        selected_product = None
    else:
        selected_product = st.selectbox("商品を選択してください", cand_df["商品名"])
else:
    selected_product = None

# 2. 商品が選択されたら入力画面
if selected_product:
    now = datetime.now()
    sheet_name = f"{now.month}月パン廃棄"
    last_day = (now.replace(day=28) + timedelta(days=4)).replace(day=1) - timedelta(days=1)

    # Excelファイルとシート準備
    if os.path.exists(DATA_FILE):
        wb = load_workbook(DATA_FILE)
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.create_sheet(sheet_name)
            ws.cell(row=1, column=1, value="商品名")
            for i in range(last_day.day):
                ws.cell(row=1, column=i + 2, value=f"{i+1:02d}")
            ws.cell(row=1, column=last_day.day + 2, value="ロス合計個数")
            set_column_width(ws, last_day.day)
            wb.save(DATA_FILE)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        ws.cell(row=1, column=1, value="商品名")
        for i in range(last_day.day):
            ws.cell(row=1, column=i + 2, value=f"{i+1:02d}")
        ws.cell(row=1, column=last_day.day + 2, value="ロス合計個数")
        set_column_width(ws, last_day.day)
        wb.save(DATA_FILE)

    # 既存行検索
    names = [ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1) if ws.cell(row=r, column=1).value]
    if selected_product in names:
        row_idx = names.index(selected_product) + 2
    else:
        row_idx = ws.max_row + 1
        ws.cell(row=row_idx, column=1, value=selected_product)

    # 今日の列（1日→列2）
    col_idx = now.day + 1

    prev_qty = ws.cell(row=row_idx, column=col_idx).value or 0

    qty = st.number_input(f"{selected_product} の廃棄個数を入力", min_value=0, step=1, value=0)
    st.info(f"本日すでに記録されている数量: {prev_qty} 個")
    add_mode = st.checkbox("既存の数に合計する")

    if st.button("記録"):
        try:
            new_qty = prev_qty + qty if add_mode else qty
            ws.cell(row=row_idx, column=col_idx, value=new_qty)

            # 合計列更新
            total_loss = sum(ws.cell(row=row_idx, column=i + 2).value or 0 for i in range(last_day.day))
            ws.cell(row=row_idx, column=last_day.day + 2, value=total_loss)

            wb.save(DATA_FILE)
            st.success(f"{selected_product} の廃棄 {new_qty} 個 を記録しました")
        except Exception as e:
            st.error(f"保存に失敗: {e}")

    # 今月シートを表示
    st.subheader("📊 今月の廃棄データ")
    df_month = pd.read_excel(DATA_FILE, sheet_name=sheet_name, engine="openpyxl")
    st.dataframe(df_month)

# 終了ボタン
if st.button("入力を終了する"):
    st.warning("✅ 入力を終了しました。画面を閉じてください。")
    st.stop()
